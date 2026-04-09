import csv
import os
from datetime import datetime, timedelta

from modules.empleados import buscar_empleado
from modules.ventas import cargar_ventas

FORMATO_FECHA_VENTA = "%Y-%m-%d %H:%M:%S"
FORMATO_FECHA_UI = "%Y-%m-%d"

TIPO_DIARIO = "diario"
TIPO_SEMANAL = "semanal"
TIPO_MENSUAL = "mensual"

TIPOS_REPORTE = {
    TIPO_DIARIO: "Diario",
    TIPO_SEMANAL: "Semanal",
    TIPO_MENSUAL: "Mensual",
}


def excel_disponible():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return False
    return True


def listar_tipos_reporte():
    return [TIPOS_REPORTE[tipo] for tipo in [TIPO_DIARIO, TIPO_SEMANAL, TIPO_MENSUAL]]


def normalizar_tipo_reporte(tipo_reporte):
    texto = str(tipo_reporte or "").strip().lower()
    if texto in TIPOS_REPORTE:
        return texto

    for clave, etiqueta in TIPOS_REPORTE.items():
        if texto == etiqueta.lower():
            return clave

    raise ValueError("Tipo de reporte inválido. Usá Diario, Semanal o Mensual.")


def _parsear_fecha_base(fecha_base):
    if fecha_base is None:
        return datetime.now()

    if isinstance(fecha_base, datetime):
        return fecha_base

    texto = str(fecha_base).strip()
    if not texto:
        return datetime.now()

    try:
        return datetime.strptime(texto, FORMATO_FECHA_UI)
    except ValueError as exc:
        raise ValueError("La fecha debe tener formato YYYY-MM-DD.") from exc


def _rango_periodo(tipo_reporte, fecha_base):
    base = _parsear_fecha_base(fecha_base)
    inicio = datetime(base.year, base.month, base.day)

    if tipo_reporte == TIPO_DIARIO:
        fin = inicio + timedelta(days=1)
        etiqueta = f"{inicio.strftime('%d/%m/%Y')} ({TIPOS_REPORTE[tipo_reporte]})"
        return inicio, fin, etiqueta

    if tipo_reporte == TIPO_SEMANAL:
        inicio = inicio - timedelta(days=inicio.weekday())
        fin = inicio + timedelta(days=7)
        etiqueta = (
            f"Semana del {inicio.strftime('%d/%m/%Y')} "
            f"al {(fin - timedelta(days=1)).strftime('%d/%m/%Y')}"
        )
        return inicio, fin, etiqueta

    if tipo_reporte == TIPO_MENSUAL:
        inicio = datetime(base.year, base.month, 1)
        if base.month == 12:
            fin = datetime(base.year + 1, 1, 1)
        else:
            fin = datetime(base.year, base.month + 1, 1)
        etiqueta = inicio.strftime("%m/%Y")
        return inicio, fin, etiqueta

    raise ValueError("Tipo de reporte inválido. Usá Diario, Semanal o Mensual.")


def _parsear_fecha_venta(venta):
    return datetime.strptime(venta.get("fecha", ""), FORMATO_FECHA_VENTA)


def _nombre_empleado(empleado_id):
    empleado = buscar_empleado(empleado_id)
    if empleado:
        return empleado.get("nombre") or f"ID {empleado_id}"
    return f"ID {empleado_id}"


def _filas_resumen_exportacion(reporte):
    return [
        ["Campo", "Valor"],
        ["Periodo", reporte["periodo_label"]],
        ["Tipo de reporte", reporte["etiqueta_tipo"]],
        ["Fecha base", reporte["fecha_base"]],
        ["Rango desde", reporte["inicio"]],
        ["Rango hasta", reporte["fin"]],
        ["Cantidad de ventas", reporte["cantidad_ventas"]],
        ["Total facturado", reporte["total_facturado"]],
        ["Ticket promedio", reporte["ticket_promedio"]],
        ["Unidades vendidas", reporte["unidades_vendidas"]],
        ["Productos distintos", reporte["productos_distintos"]],
    ]


def _exportar_reporte_csv(reporte, ruta_destino):
    with open(ruta_destino, "w", newline="", encoding="utf-8-sig") as archivo:
        writer = csv.writer(archivo)

        writer.writerow(["Resumen"])
        writer.writerows(_filas_resumen_exportacion(reporte))
        writer.writerow([])

        writer.writerow(["Productos mas vendidos"])
        writer.writerow(["Producto", "Cantidad vendida", "Facturado", "Tickets"])
        for producto in reporte["productos_mas_vendidos"]:
            writer.writerow(
                [
                    producto["nombre"],
                    producto["cantidad_vendida"],
                    producto["facturado"],
                    producto["apariciones"],
                ]
            )
        writer.writerow([])

        writer.writerow(["Formas de pago"])
        writer.writerow(["Forma", "Operaciones", "Monto total"])
        for pago in reporte["formas_pago"]:
            writer.writerow(
                [
                    pago["forma_pago"],
                    pago["cantidad_operaciones"],
                    pago["monto_total"],
                ]
            )
        writer.writerow([])

        writer.writerow(["Desempeño por empleado"])
        writer.writerow(["Empleado", "Ventas", "Unidades", "Facturado"])
        for empleado in reporte["empleados_destacados"]:
            writer.writerow(
                [
                    empleado["nombre"],
                    empleado["cantidad_ventas"],
                    empleado["unidades_vendidas"],
                    empleado["monto_total"],
                ]
            )
        writer.writerow([])

        writer.writerow(["Ventas por fecha"])
        writer.writerow(["Fecha", "Ventas", "Total"])
        for dia in reporte["ventas_por_fecha"]:
            writer.writerow([dia["fecha"], dia["ventas"], dia["total"]])


def _exportar_reporte_xlsx(reporte, ruta_destino):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("No está disponible la exportación directa a Excel (.xlsx).") from exc

    wb = Workbook()

    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    for fila in _filas_resumen_exportacion(reporte):
        ws_resumen.append(fila)

    ws_productos = wb.create_sheet("Productos")
    ws_productos.append(["Producto", "Cantidad vendida", "Facturado", "Tickets"])
    for producto in reporte["productos_mas_vendidos"]:
        ws_productos.append(
            [
                producto["nombre"],
                producto["cantidad_vendida"],
                producto["facturado"],
                producto["apariciones"],
            ]
        )

    ws_pagos = wb.create_sheet("Pagos")
    ws_pagos.append(["Forma", "Operaciones", "Monto total"])
    for pago in reporte["formas_pago"]:
        ws_pagos.append([pago["forma_pago"], pago["cantidad_operaciones"], pago["monto_total"]])

    ws_empleados = wb.create_sheet("Empleados")
    ws_empleados.append(["Empleado", "Ventas", "Unidades", "Facturado"])
    for empleado in reporte["empleados_destacados"]:
        ws_empleados.append(
            [
                empleado["nombre"],
                empleado["cantidad_ventas"],
                empleado["unidades_vendidas"],
                empleado["monto_total"],
            ]
        )

    ws_dias = wb.create_sheet("Por fecha")
    ws_dias.append(["Fecha", "Ventas", "Total"])
    for dia in reporte["ventas_por_fecha"]:
        ws_dias.append([dia["fecha"], dia["ventas"], dia["total"]])

    wb.save(ruta_destino)


def exportar_reporte_ventas(reporte, ruta_destino):
    extension = os.path.splitext(str(ruta_destino or "").strip())[1].lower()
    if extension == ".xlsx":
        _exportar_reporte_xlsx(reporte, ruta_destino)
        return ruta_destino

    _exportar_reporte_csv(reporte, ruta_destino)
    return ruta_destino


def generar_reporte_ventas(tipo_reporte="diario", fecha_base=None):
    tipo = normalizar_tipo_reporte(tipo_reporte)
    inicio, fin, etiqueta_periodo = _rango_periodo(tipo, fecha_base)

    ventas_filtradas = []
    for venta in cargar_ventas():
        try:
            fecha_venta = _parsear_fecha_venta(venta)
        except ValueError:
            continue

        if inicio <= fecha_venta < fin:
            ventas_filtradas.append((fecha_venta, venta))

    total_facturado = 0.0
    unidades_vendidas = 0
    productos = {}
    formas_pago = {}
    empleados = {}
    ventas_por_fecha = {}

    for fecha_venta, venta in ventas_filtradas:
        total_venta = float(venta.get("total", 0) or 0)
        total_facturado += total_venta

        fecha_clave = fecha_venta.strftime(FORMATO_FECHA_UI)
        resumen_fecha = ventas_por_fecha.setdefault(
            fecha_clave,
            {"fecha": fecha_clave, "ventas": 0, "total": 0.0},
        )
        resumen_fecha["ventas"] += 1
        resumen_fecha["total"] += total_venta

        forma_pago = str(venta.get("forma_pago", "sin dato")).lower()
        resumen_pago = formas_pago.setdefault(
            forma_pago,
            {"forma_pago": forma_pago, "cantidad_operaciones": 0, "monto_total": 0.0},
        )
        resumen_pago["cantidad_operaciones"] += 1
        resumen_pago["monto_total"] += total_venta

        empleado_id = str(venta.get("empleado_id", ""))
        resumen_empleado = empleados.setdefault(
            empleado_id,
            {
                "empleado_id": empleado_id,
                "nombre": _nombre_empleado(empleado_id),
                "cantidad_ventas": 0,
                "unidades_vendidas": 0,
                "monto_total": 0.0,
            },
        )
        resumen_empleado["cantidad_ventas"] += 1
        resumen_empleado["monto_total"] += total_venta

        for producto in venta.get("productos", []):
            producto_id = str(producto.get("id", ""))
            cantidad = int(producto.get("cantidad", 0) or 0)
            subtotal = float(producto.get("subtotal", 0) or 0)
            unidades_vendidas += cantidad
            resumen_empleado["unidades_vendidas"] += cantidad

            resumen_producto = productos.setdefault(
                producto_id,
                {
                    "id": producto_id,
                    "nombre": producto.get("nombre", f"Producto {producto_id}"),
                    "cantidad_vendida": 0,
                    "facturado": 0.0,
                    "apariciones": 0,
                },
            )
            resumen_producto["cantidad_vendida"] += cantidad
            resumen_producto["facturado"] += subtotal
            resumen_producto["apariciones"] += 1

    total_facturado = round(total_facturado, 2)
    cantidad_ventas = len(ventas_filtradas)
    ticket_promedio = round(total_facturado / cantidad_ventas, 2) if cantidad_ventas else 0.0

    productos_ordenados = sorted(
        productos.values(),
        key=lambda producto: (producto["cantidad_vendida"], producto["facturado"], producto["nombre"]),
        reverse=True,
    )
    empleados_ordenados = sorted(
        empleados.values(),
        key=lambda empleado: (empleado["monto_total"], empleado["cantidad_ventas"], empleado["nombre"]),
        reverse=True,
    )
    pagos_ordenados = sorted(
        formas_pago.values(),
        key=lambda pago: (pago["monto_total"], pago["cantidad_operaciones"], pago["forma_pago"]),
        reverse=True,
    )
    dias_ordenados = sorted(
        ventas_por_fecha.values(),
        key=lambda dia: dia["fecha"],
    )

    for coleccion in [productos_ordenados, empleados_ordenados, pagos_ordenados, dias_ordenados]:
        for item in coleccion:
            for clave, valor in list(item.items()):
                if isinstance(valor, float):
                    item[clave] = round(valor, 2)

    return {
        "tipo_reporte": tipo,
        "etiqueta_tipo": TIPOS_REPORTE[tipo],
        "fecha_base": inicio.strftime(FORMATO_FECHA_UI),
        "periodo_label": etiqueta_periodo,
        "inicio": inicio.strftime(FORMATO_FECHA_VENTA),
        "fin": (fin - timedelta(seconds=1)).strftime(FORMATO_FECHA_VENTA),
        "cantidad_ventas": cantidad_ventas,
        "total_facturado": total_facturado,
        "ticket_promedio": ticket_promedio,
        "unidades_vendidas": unidades_vendidas,
        "productos_distintos": len(productos_ordenados),
        "productos_mas_vendidos": productos_ordenados,
        "empleados_destacados": empleados_ordenados,
        "formas_pago": pagos_ordenados,
        "ventas_por_fecha": dias_ordenados,
        "mejor_producto": productos_ordenados[0] if productos_ordenados else None,
        "mejor_empleado": empleados_ordenados[0] if empleados_ordenados else None,
        "mejor_dia": max(dias_ordenados, key=lambda dia: dia["total"], default=None),
    }
